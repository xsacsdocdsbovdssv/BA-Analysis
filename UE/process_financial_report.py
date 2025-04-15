import pandas as pd
import os
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import datetime


def clean_numeric_row(row):
    return pd.to_numeric(row.astype(str).str.replace(",", "").str.strip(), errors='coerce')

def process_financial_report(input_file:str) -> str:
    """处理半天妖财务报表的主函数"""

    # 1. 准备文件路径
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M')
    output_file = input_file.replace(".xlsx", f"_自动处理_{timestamp}.xlsx")
    temp_file = input_file.replace(".xlsx", "_temp.xlsx")

    try:
        # 2. 读取并处理原始数据
        print("正在读取原始数据...")
        df_raw = pd.read_excel(input_file, header=1)
        df_items = df_raw.set_index("项 目")
        df_items = df_items.drop(["行次", "合计"], axis=1, errors="ignore")
        df_items = df_items.apply(clean_numeric_row)
        # 清洗为数字
        df_items = df_items.apply(lambda row: pd.to_numeric(row.astype(str).str.replace(",", "").str.strip(), errors='coerce'))

        # 3. 计算关键指标
        print("正在计算经营指标...")
        number_of_stores = df_items.shape[1]

        # 收入类计算
        Actual_revenue = df_items.loc["二、主营业务收入"] - df_items.loc["23.外卖平台手续费"]
        dine_in = df_items.loc["1.堂食实收收入"]
        takeout = df_items.loc["2.外卖实收收入（含外卖手续费）"] - df_items.loc["23.外卖平台手续费"]

        # 成本类计算
        Ingredients_cost = df_items.loc["减：营业成本"]
        Packaging_cost = df_items.loc["减：包装成本"]
        Rent = df_items.loc["1.租金"]
        Labor = (df_items.loc["2.工资"] + df_items.loc["3.员工福利"] +
                 df_items.loc["4.考核工资"] + df_items.loc["11.督导均摊费"])

        # 费用类计算
        Electricity = df_items.loc["5.电费"]
        Other_platform_fees = df_items.loc["24.其他平台手续费"]
        Other_miscellaneous_fees = (
                df_items.loc["6.水费"] + df_items.loc["7.前厅低值易耗品"] +
                df_items.loc["8.后厨低值易耗品"] + df_items.loc["9.电话费\宽带费"] +
                df_items.loc["10.交通费\住宿费"] + df_items.loc["12.运费"] +
                df_items.loc["13.广告费"] + df_items.loc["14.维修费"] +
                df_items.loc["15.消杀费\卫生费"] + df_items.loc["16.医疗用品"] +
                df_items.loc["17.拓展费"] + df_items.loc["18.日常资产"] +
                df_items.loc["19.装修改造费"] + df_items.loc["21.账户管理费"] +
                df_items.loc["22.其他费用"] + df_items.loc["25.菜金折扣\现金券消费"]
        )

        # 其他计算
        Headquarters_fees = (df_items.loc["26.服务费"] +
                             df_items.loc["27.品牌使用费"] +
                             df_items.loc["28.美团广告抽点"])
        Taxes = (df_items.loc["20.税费"] +
                 df_items.loc["减：税金及附加"] +
                 df_items.loc["减：企业所得税费用"])
        Net_profit = df_items.loc["八、净利润"]
        Other_income = (df_items.loc["加：营业外收入"] +
                        df_items.loc["加：其他业务收入"])
        Takeout_platform_fees = df_items.loc["23.外卖平台手续费"]
        # Actual_labor = df_items.loc["实际人工"]
        # Adjusted_net_profit = df_items.loc["调整后净利润"]
        # daily_avg_tables = df_items.loc["日均开台数"]
        # daily_avg_takeouts = df_items.loc["日均外卖单数"]
        # avg_price_per_table = df_items.loc["台均价"]
        # avg_price_per_takeout = df_items.loc["外卖单均价"]
        # avg_food_cost_per_order = df_items.loc["单均食材成本"]
        # avg_packaging_cost_per_order = df_items.loc["单均外卖包装成本"]

        # 4. 创建结果DataFrame
        df_calc = pd.DataFrame({
            "门店数量": number_of_stores,
            "实收": Actual_revenue,
            "堂食实收": dine_in,
            "外卖实收": takeout,
            "食材成本": Ingredients_cost,
            "包装成本": Packaging_cost,
            "租金": Rent,
            "人工": Labor,
            "电费": Electricity,
            "其他平台费": Other_platform_fees,
            "其他杂费": Other_miscellaneous_fees,
            "总部收取费用": Headquarters_fees,
            "税费": Taxes,
            "净利润": Net_profit,
            "其他收入": Other_income,
            "外卖平台费": Takeout_platform_fees,
            "实际人工": np.nan,
            "调整后净利润": np.nan,
            "日均开台数": np.nan,
            "日均外卖单数": np.nan,
            "台均价": np.nan,
            "外卖单均价": np.nan,
            "单均食材成本": np.nan,
            "单均外卖包装成本": np.nan

        })
        # df_calc["实际人工"] = Actual_labor
        # df_calc["调整后净利润"] = Adjusted_net_profit
        # df_calc["日均开台数"] = daily_avg_tables
        # df_calc["日均外卖单数"] = daily_avg_takeouts
        # df_calc["台均价"] = avg_price_per_table
        # df_calc["外卖单均价"] = avg_price_per_takeout
        # df_calc["单均食材成本"] = avg_food_cost_per_order
        # df_calc["单均外卖包装成本"] = avg_packaging_cost_per_order


        # 5. 写入临时Excel文件
        print("正在生成临时文件...")
        df_calc.to_excel(temp_file, sheet_name="经营分析", index=True, index_label="项 目")

        # 6. 格式处理
        print("正在格式化输出文件...")
        wb = load_workbook(temp_file)
        ws = wb["经营分析"]

        # 设置透明填充样式
        transparent_fill = PatternFill(fill_type=None)

        # 清除所有单元格格式
        for row in ws.iter_rows():
            for cell in row:
                cell.font = None
                cell.border = None
                cell.fill = transparent_fill
                cell.alignment = None
                cell.number_format = 'General'

        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # 7. 保存最终文件
        wb.save(output_file)
        print(f"文件已成功保存为：{output_file}")

        return output_file

    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")
        raise
    finally:
        # 清理临时文件
        if os.path.exists(temp_file):
            os.remove(temp_file)
            print("已清理临时文件")
